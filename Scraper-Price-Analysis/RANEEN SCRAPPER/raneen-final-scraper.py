from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, ElementClickInterceptedException, StaleElementReferenceException
)
import time
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# === Chrome Setup ===
options = Options()
# options.add_argument('--headless=new')  # Uncomment for headless mode
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1920,1080')
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)

# === Input Excel ===
input_excel = "raneen-target-links.xlsx"
df = pd.read_excel(input_excel, header=1)
df.columns = df.columns.str.strip()
df = df.dropna(subset=["Category", "URL"])
category_links = list(zip(df["Category"], df["URL"]))

# === Output Directory ===
output_dir = "raneen-products"
os.makedirs(output_dir, exist_ok=True)

# === Excel Styling ===
def style_excel_file(path):
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill(start_color="990000", end_color="990000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="000000"))

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = body_font

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    wb.save(path)

# === Helpers ===
def normalize_price(text):
    if not text:
        return None
    return int(re.sub(r"[^\d]", "", text))

def extract_sku(name):
    if not name:
        return ""
    upper_name = name.upper()
    pattern = r'\b(?:[A-Z0-9]{1,10})(?:[ .\-_]{0,1}[A-Z0-9]{1,10}){0,2}\b'
    matches = re.findall(pattern, upper_name)
    for candidate in reversed(matches):
        if any(c.isalpha() for c in candidate) and len(candidate.replace(" ", "")) >= 2:
            return candidate.strip()
    return ""

def normalize_sku(sku):
    return re.sub(r'[-_\s]', '', sku).lower() if sku else ""

# === Start Scraping ===
print("🚀 Starting Raneen Scraper")
for category, url in category_links:
    print(f"\n➡️ Scraping Category: {category} | URL: {url}")
    driver.get(url)
    time.sleep(2)

    # Load all products using "Load More"
    prev_count = -1
    same_count_repeats = 0
    max_repeats = 3

    while same_count_repeats < max_repeats:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        products = driver.find_elements(By.CSS_SELECTOR, "div.product-item-info")
        current_count = len(products)
        print(f"🔍 Products loaded so far: {current_count}")
        if current_count == prev_count:
            same_count_repeats += 1
        else:
            same_count_repeats = 0
            prev_count = current_count

    product_cards = driver.find_elements(By.CSS_SELECTOR, "div.product-item-info")
    print(f"✅ Total products loaded: {len(product_cards)}")

    data = []
    for card in product_cards:
        try:
            title_el = card.find_element(By.CSS_SELECTOR, "a.product-item-link")
            title = title_el.text.strip()
            product_url = title_el.get_attribute("href")

            # === Final Unified Price Logic (Handles all formats) ===
            try:
                price_box = card.find_element(By.CSS_SELECTOR, ".price-box.price-final_price")

                new_price = old_price = None

                # Case 1: Discounted (special + old)
                special_price_els = price_box.find_elements(By.CSS_SELECTOR, ".special-price .price-wrapper")
                old_price_els = price_box.find_elements(By.CSS_SELECTOR, ".old-price .price-wrapper")

                if special_price_els:
                    new_price = normalize_price(special_price_els[0].text)
                    if old_price_els:
                        old_price = normalize_price(old_price_els[0].text)
                else:
                    # Case 2: Regular price with wrapper
                    regular_price_els = price_box.find_elements(By.CSS_SELECTOR, ".price-container .price-wrapper")
                    if regular_price_els:
                        new_price = normalize_price(regular_price_els[0].text)
                    else:
                        # Case 3: Raw text fallback (no .price-wrapper)
                        current_price_span = price_box.find_elements(By.CSS_SELECTOR, ".current-price")
                        old_price_span = price_box.find_elements(By.CSS_SELECTOR, ".old-price")

                        if current_price_span:
                            new_price = normalize_price(current_price_span[0].text)
                        if old_price_span:
                            old_price = normalize_price(old_price_span[0].text)

            except Exception as e:
                print("⚠️ Failed to extract price:", e)
                new_price = None
                old_price = None

            product_code = extract_sku(title)
            normalized_code = normalize_sku(product_code)

            data.append({
                "Item Name": title,
                "Old Price": old_price,
                "New Price": new_price,
                "Product URL": product_url,
                "Product Code": product_code,
                "Normalized Code": normalized_code
            })

        except Exception as e:
            print("⚠️ Skipping product due to error:", e)

    # Save to Excel
    if data:
        timestamp = datetime.now().strftime("%Y-%m-%d")
        safe_category = re.sub(r"[^\w\s-]", "", category).replace(" ", "_")
        output_file = os.path.join(output_dir, f"raneen_{safe_category}_{timestamp}.xlsx")
        df_out = pd.DataFrame(data)
        df_out.to_excel(output_file, index=False, engine="openpyxl")
        style_excel_file(output_file)
        print(f"💾 Saved {len(data)} products to {output_file}")
    else:
        print("⚠️ No product data extracted.")

# === Done ===
driver.quit()
print("🏁 All categories processed for Raneen.")
print("✅ Scraping completed successfully!")
print(f"📂 Output files saved in: {output_dir}")