from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    ElementClickInterceptedException, TimeoutException, StaleElementReferenceException
)
import time
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# === Chrome Setup ===
options = Options()
options.add_argument('--headless=new')
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1920,1080')
options.add_argument('--lang=ar')
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)

# === Input Excel ===
input_excel = "btech-target-links.xlsx"
df = pd.read_excel(input_excel, header=1)
df.columns = df.columns.str.strip()
df = df.dropna(subset=["Category", "URL"])
category_links = list(zip(df["Category"], df["URL"]))

# === Output Directory ===
output_dir = "btech-products"
os.makedirs(output_dir, exist_ok=True)

# === Excel Styling ===
def style_excel_file(path):
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="000000"))

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = body_font

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    wb.save(path)

# === Helpers ===
def normalize_price(price_text):
    return int(price_text.replace(",", "").strip()) if price_text else None

def extract_sku(name):
    if not name:
        return ""

    name = name.upper().replace("\u200f", "")  # remove RTL char

    # Regex: match blocks of 3+ alphanum (optionally separated by space, dash, plus), at end or after dash
    pattern = r'(?:-\s*)?([A-Z0-9][A-Z0-9\s\+\-]{2,})$'
    matches = re.findall(pattern, name)
    # Fallback: also match any block of 3+ alphanum (with optional spaces/pluses/dashes)
    if not matches:
        pattern2 = r'([A-Z0-9][A-Z0-9\s\+\-]{2,})'
        matches = re.findall(pattern2, name)
    # Filter: must have at least 1 letter and at least 3 chars
    candidates = [m.strip() for m in matches if any(c.isalpha() for c in m) and len(m.strip()) >= 3]
    return candidates[-1] if candidates else ""

def normalize_sku(sku):
    return re.sub(r'[^a-zA-Z0-9]', '', sku).lower() if sku else ""

def extract_total_expected_products(driver):
    try:
        el = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "product-search-item-count"))
        )
        return int(el.text.strip())
    except:
        print("⚠️ Could not extract expected count.")
        return None

# === Scrape Each Category ===
print("🚀 Starting Btech Scraper")
for category, url in category_links:
    print(f"\n➡️ Category: {category} | URL: {url}")
    driver.get(url)
    time.sleep(2)

    safe_category = re.sub(r"[^\w\s-]", "", category).replace(" ", "_")
    date_str = datetime.now().strftime("%Y-%m-%d")
    output_path = os.path.join(output_dir, f"btech_{safe_category}_{date_str}.xlsx")

    expected_total = extract_total_expected_products(driver)
    max_scrape_limit = expected_total + 2 if expected_total else float("inf")
    print(f"📊 Expected products: {expected_total} | Max scrape: {max_scrape_limit}")

    previous_count = -1
    attempt = 0
    max_attempts = 40

    while attempt < max_attempts:
        time.sleep(2)
        products = driver.find_elements(By.CSS_SELECTOR, "div.plpContentWrapper")
        current_count = len(products)
        print(f"🟨 Products loaded: {current_count}")

        if expected_total and current_count >= expected_total + 2:
            print("🛑 Expected count reached.")
            break
        if current_count == previous_count:
            print("✅ No new products loaded.")
            break
        previous_count = current_count

        try:
            load_more_btn = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.amscroll-load-button"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", load_more_btn)
            time.sleep(1)
            try:
                load_more_btn.click()
                print("🔁 Clicked Load More")
            except ElementClickInterceptedException:
                print("⚠️ Intercepted, retrying with JS")
                driver.execute_script("arguments[0].click();", load_more_btn)
        except TimeoutException:
            print("ℹ️ Load More not found.")
            break
        attempt += 1

    # === Parse Products ===
    data = []
    wrappers = driver.find_elements(By.CSS_SELECTOR, "a.listingWrapperSection")

    for wrapper in wrappers:
        try:
            title_el = wrapper.find_elements(By.CSS_SELECTOR, "h2.plpTitle")
            if not title_el or not title_el[0].text.strip():
                continue
            title = title_el[0].text.strip()
            new_price_el = wrapper.find_elements(By.CSS_SELECTOR, "span.special-price span.price-wrapper")
            old_price_el = wrapper.find_elements(By.CSS_SELECTOR, "span.old-price.was-price span.price-wrapper")

            new_price = normalize_price(new_price_el[0].text) if new_price_el else None
            old_price = normalize_price(old_price_el[0].text) if old_price_el else None

            product_url = wrapper.get_attribute("href")
            product_code = extract_sku(title)
            normalized_code = normalize_sku(product_code)

            data.append({
                "Item Name": title,
                "Old Price": old_price,
                "New Price": new_price,
                "Product URL": product_url,
                "Product Code": product_code,
                "Normalized Code": normalized_code
            })
        except Exception as e:
            print("❌ Skipped product:", e)

    # === Save Output ===
    if data:
        df_out = pd.DataFrame(data)
        df_out.to_excel(output_path, index=False, engine='openpyxl')
        style_excel_file(output_path)
        print(f"✅ Saved {len(data)} products to {output_path}")
    else:
        print("⚠️ No data extracted.")

# === Done ===
driver.quit()
print("🏁 All categories processed for Btech.")
