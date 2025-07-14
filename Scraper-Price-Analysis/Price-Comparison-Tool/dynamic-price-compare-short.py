import os
import re
import pandas as pd
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))

RETAILER_FOLDERS = {
    "2B": os.path.join(BASE_DIR, "2B SCRAPPER", "2b-Products"),
    "BTECH": os.path.join(BASE_DIR, "BTECH SCRAPPER", "Btech-Products"),
    "RANEEN": os.path.join(BASE_DIR, "RANEEN SCRAPPER", "Raneen-Products"),
}

REQUIRED_COLUMNS = ["Item Name", "New Price", "Normalized Code", "Product URL"]
CONFIDENCE_THRESHOLD = 10
HIGHLIGHT_THRESHOLD = 30

OUTPUT_FOLDER = os.path.join(BASE_DIR, "Price-Comparison-Results", "short")
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def log(msg): print(f"[LOG] {msg}")

def extract_info_from_filename(filename):
    match = re.match(r"([a-zA-Z0-9]+)_([A-Za-z0-9\-]+)_\d{4}-\d{2}-\d{2}\.xlsx", filename)
    if match:
        return match.group(1).upper(), match.group(2)
    return None, None

def compute_confidence(names):
    if len(names) < 2: return 0.0
    pairs = [(names[i], names[j]) for i in range(len(names)) for j in range(i + 1, len(names))]
    scores = [SequenceMatcher(None, a, b).ratio() for a, b in pairs]
    return round(sum(scores) / len(scores) * 100, 1)

def export_results(df_rows, filename):
    if not df_rows:
        log(f"⚠️ No data to export for {filename}")
        return

    df = pd.DataFrame(df_rows).sort_values(by="Confidence", ascending=False)
    out_path = os.path.join(OUTPUT_FOLDER, filename)
    df.to_excel(out_path, index=False, engine="openpyxl")

    wb = load_workbook(out_path)
    ws = wb.active
    align_center = Alignment(horizontal="center", vertical="center")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    headers = df.columns.tolist()
    conf_idx = headers.index("Confidence")
    price_idx = headers.index("Best Price")
    retailer_idx = headers.index("Lowest Retailer")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_len + 2
        for cell in col:
            cell.alignment = align_center

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        try:
            if float(row[conf_idx].value) < HIGHLIGHT_THRESHOLD:
                for cell in row:
                    cell.fill = yellow_fill
            row[price_idx].font = bold_font
            row[retailer_idx].font = bold_font
        except:
            continue

    wb.save(out_path)
    log(f"✅ Exported: {out_path}")

# === Main script ===
log("🔍 Scanning retailer folders...")

category_map = {}
matched_categories = 0
skipped_categories = 0

for retailer, folder_path in RETAILER_FOLDERS.items():
    log(f"🔎 Checking: {folder_path}")
    if not os.path.exists(folder_path):
        log(f"❌ Folder not found: {folder_path}")
        continue

    files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
    log(f"📁 {retailer} has {len(files)} file(s)")

    for filename in files:
        r_name, category = extract_info_from_filename(filename)
        if r_name and category:
            category_map.setdefault(category, {}).setdefault(retailer, os.path.join(folder_path, filename))

log(f"📦 Found {len(category_map)} categories.")
for cat, data in category_map.items():
    log(f"  - {cat}: {list(data.keys())}")

for category, sources in category_map.items():
    if len(sources) < 2:
        log(f"⏭️ Skipping '{category}' (only {len(sources)} source(s))")
        skipped_categories += 1
        continue

    matched_categories += 1
    log(f"\n🚀 Processing category: {category}")
    all_data = []

    for retailer, file_path in sources.items():
        log(f"📥 Reading file for {retailer}: {file_path}")
        try:
            df = pd.read_excel(file_path)
            df.columns = df.columns.str.strip()

            if not all(col in df.columns for col in REQUIRED_COLUMNS):
                log(f"⚠️ Skipping {retailer} in {category} — missing required columns")
                continue

            df = df[REQUIRED_COLUMNS].copy()
            df["New Price"] = pd.to_numeric(df["New Price"], errors="coerce")
            df["Normalized Code"] = df["Normalized Code"].astype(str).str.lower()
            df["Source"] = retailer
            all_data.append(df)

        except Exception as e:
            log(f"❌ Error reading {file_path}: {e}")
            continue

    if len(all_data) < 2:
        log(f"⚠️ Not enough valid files to compare for {category}")
        continue

    combined = pd.concat(all_data, ignore_index=True)
    matched, unmatched = [], []

    for code, group in combined.groupby("Normalized Code"):
        group = group.dropna(subset=["New Price"])
        if len(group) < 2:
            continue

        names = group["Item Name"].tolist()
        confidence = compute_confidence(names)
        best_idx = group["New Price"].idxmin()
        best_row = group.loc[best_idx]

        result = {
            "Item Name": best_row["Item Name"],
            "Normalized Code": code,
            "Confidence": confidence,
            "Best Price": best_row["New Price"],
            "Lowest Retailer": best_row["Source"],
            "Product URL": best_row["Product URL"]
        }

        if confidence >= CONFIDENCE_THRESHOLD:
            matched.append(result)
        else:
            unmatched.append(result)

    log(f"📊 {category}: Matched = {len(matched)}, Unmatched = {len(unmatched)}")
    export_results(matched, f"cross-compare-{category}-short-matched.xlsx")
    export_results(unmatched, f"cross-compare-{category}-short-unmatched.xlsx")

log(f"\n✅ All comparisons completed.")
log(f"📁 Matched categories: {matched_categories}")
log(f"📁 Skipped categories: {skipped_categories}")

skipped_list = {
    cat: list(srcs.keys())
    for cat, srcs in category_map.items()
    if len(srcs) < 2
}

if skipped_list:
    log("\n🚫 Skipped Categories (appear in only one retailer):")
    for cat, retailers in skipped_list.items():
        log(f"  - {cat} (from: {', '.join(retailers)})")
else:
    log("✅ No skipped categories due to missing retailer coverage.")
