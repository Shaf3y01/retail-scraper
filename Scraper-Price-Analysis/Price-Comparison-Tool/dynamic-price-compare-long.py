import os
import re
import pandas as pd
from datetime import datetime
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# === Setup ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))

RETAILER_FOLDERS = {
    "2B": os.path.join(BASE_DIR, "2B SCRAPPER", "2b-Products"),
    "Btech": os.path.join(BASE_DIR, "BTECH SCRAPPER", "Btech-Products"),
    "Raneen": os.path.join(BASE_DIR, "RANEEN SCRAPPER", "Raneen-Products"),
}

OUTPUT_FOLDER = os.path.join(BASE_DIR, "Price-Comparison-Results", "long")
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

REQUIRED_COLUMNS = ["Item Name", "New Price", "Normalized Code"]
HIGHLIGHT_CONFIDENCE_WEAK = 30
HIGHLIGHT_CONFIDENCE_UNMATCHED = 10

# === Logging ===
def log(msg):
    print(f"[LOG] {msg}")

# === Utilities ===
def extract_info_from_filename(filename):
    match = re.match(r"([a-zA-Z0-9]+)_([A-Za-z0-9\-]+)_\d{4}-\d{2}-\d{2}\.xlsx", filename)
    if match:
        return match.group(1).capitalize(), match.group(2)
    return None, None

def match_score(a, b):
    if pd.isna(a) or pd.isna(b):
        return 0.0
    return SequenceMatcher(None, str(a), str(b)).ratio()

def prepare(df, retailer):
    df.columns = df.columns.str.strip()
    if not all(col in df.columns for col in REQUIRED_COLUMNS):
        return None
    df = df[REQUIRED_COLUMNS].copy()
    df["Retailer"] = retailer
    df = df.dropna(subset=["Normalized Code"])
    df["Normalized Code"] = df["Normalized Code"].astype(str).str.strip().str.lower()
    df["Item Name"] = df["Item Name"].astype(str).str.strip()
    df["New Price"] = pd.to_numeric(df["New Price"], errors="coerce")
    return df

def export_results(rows, filename, highlight_confidence=None):
    if not rows:
        return
    df_out = pd.concat(rows).reset_index(drop=True)
    final_cols = [
        "2B Item Name", "2B Price", "2B Item SKU",
        "Btech Item Name", "Btech Price", "Btech Item SKU",
        "Raneen Item Name", "Raneen Price", "Raneen Item SKU",
        "Confidence", "Best Price", "Lowest Retailer"
    ]
    df_out = df_out[final_cols]
    output_path = os.path.join(OUTPUT_FOLDER, filename)
    df_out.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_len + 2
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    if highlight_confidence:
        fill = PatternFill(start_color=highlight_confidence["color"], end_color=highlight_confidence["color"], fill_type="solid")
        for row in ws.iter_rows(min_row=2):
            try:
                conf = float(row[9].value)
                if conf < highlight_confidence["threshold"]:
                    for cell in row:
                        cell.fill = fill
            except:
                continue

    wb.save(output_path)
    log(f"✅ Saved to {output_path}")

# === Step 1: Scan retailer folders ===
log("🔍 Scanning retailer folders...")

category_map = {}

for retailer, folder_path in RETAILER_FOLDERS.items():
    log(f"🔎 Checking: {folder_path}")
    if not os.path.exists(folder_path):
        log(f"❌ Folder not found: {folder_path}")
        continue

    files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
    for filename in files:
        ret_name, category = extract_info_from_filename(filename)
        if ret_name and category:
            category_map.setdefault(category, {})[retailer] = os.path.join(folder_path, filename)

log(f"📦 Found {len(category_map)} categories.")
for cat, data in category_map.items():
    log(f"  - {cat}: {list(data.keys())}")

# === Step 2: Process each category ===
matched_categories = []

for category, sources in category_map.items():
    if len(sources) < 2:
        continue

    log(f"\n🚀 Processing category: {category}")
    all_dfs = []

    for retailer, path in sources.items():
        log(f"📥 Reading: {path}")
        try:
            df = pd.read_excel(path)
            prepared = prepare(df, retailer)
            if prepared is None:
                log(f"⚠️ Skipped {retailer} in {category} — missing required columns")
                continue
            all_dfs.append(prepared)
        except Exception as e:
            log(f"❌ Failed to read {retailer}: {e}")

    if len(all_dfs) < 2:
        log(f"⚠️ Not enough valid data sources for {category}")
        continue

    combined = pd.concat(all_dfs, ignore_index=True)
    grouped = combined.groupby("Normalized Code")

    matched_rows, weak_matched_rows, unmatched_rows = [], [], []

    for code, group in grouped:
        if group["Retailer"].nunique() < 2:
            continue

        merged = pd.DataFrame({"Normalized Code": [code]})
        for retailer in RETAILER_FOLDERS.keys():
            match = group[group["Retailer"] == retailer]
            if not match.empty:
                merged[f"{retailer} Item Name"] = match.iloc[0]["Item Name"]
                merged[f"{retailer} Price"] = match.iloc[0]["New Price"]
                merged[f"{retailer} Item SKU"] = code
            else:
                merged[f"{retailer} Item Name"] = "N/A"
                merged[f"{retailer} Price"] = pd.NA
                merged[f"{retailer} Item SKU"] = code

        merged["Confidence"] = 100.0
        prices = {
            r: merged[f"{r} Price"].values[0]
            for r in RETAILER_FOLDERS if pd.notna(merged[f"{r} Price"].values[0])
        }
        best_price, best_retailer = (min(prices.values()), min(prices, key=prices.get)) if prices else (None, None)
        merged["Best Price"] = best_price
        merged["Lowest Retailer"] = best_retailer

        matched_rows.append(merged)

    # Weak matches by item name
    all_seen_codes = set(grouped.groups.keys())
    remaining = combined[~combined["Normalized Code"].isin(all_seen_codes)]

    by_name = remaining.groupby("Normalized Code")
    for code, group in by_name:
        if group["Retailer"].nunique() < 2:
            continue

        merged = pd.DataFrame({"Normalized Code": [code]})
        for retailer in RETAILER_FOLDERS:
            match = group[group["Retailer"] == retailer]
            if not match.empty:
                merged[f"{retailer} Item Name"] = match.iloc[0]["Item Name"]
                merged[f"{retailer} Price"] = match.iloc[0]["New Price"]
                merged[f"{retailer} Item SKU"] = code
            else:
                merged[f"{retailer} Item Name"] = "N/A"
                merged[f"{retailer} Price"] = pd.NA
                merged[f"{retailer} Item SKU"] = code

        names = [merged.get(f"{r} Item Name").values[0] for r in RETAILER_FOLDERS]
        confidence = round(max([
            match_score(names[0], names[1]),
            match_score(names[0], names[2]),
            match_score(names[1], names[2])
        ]) * 100, 2)

        merged["Confidence"] = confidence
        prices = {
            r: merged[f"{r} Price"].values[0]
            for r in RETAILER_FOLDERS if pd.notna(merged[f"{r} Price"].values[0])
        }
        best_price, best_retailer = (min(prices.values()), min(prices, key=prices.get)) if prices else (None, None)
        merged["Best Price"] = best_price
        merged["Lowest Retailer"] = best_retailer

        if confidence >= HIGHLIGHT_CONFIDENCE_WEAK:
            weak_matched_rows.append(merged)
        else:
            unmatched_rows.append(merged)

    # Export files
    export_results(matched_rows, f"cross-compare-{category}-long-matched.xlsx")
    export_results(weak_matched_rows, f"cross-compare-{category}-long-weak-matched.xlsx", highlight_confidence={"threshold": HIGHLIGHT_CONFIDENCE_WEAK, "color": "FFFACD"})
    export_results(unmatched_rows, f"cross-compare-{category}-long-unmatched.xlsx", highlight_confidence={"threshold": HIGHLIGHT_CONFIDENCE_UNMATCHED, "color": "FF9999"})

    log(f"📊 {category}: Matched = {len(matched_rows)}, Weak = {len(weak_matched_rows)}, Unmatched = {len(unmatched_rows)}")
    matched_categories.append(category)

# === Summary ===
skipped_list = {
    cat: list(srcs.keys())
    for cat, srcs in category_map.items()
    if len(srcs) < 2
}

log("\n✅ All comparisons completed.")
log(f"📁 Matched categories: {matched_categories}")
log(f"📁 Skipped categories: {list(skipped_list.keys())}")

if skipped_list:
    log("\n🚫 Skipped Categories (appear in only one retailer):")
    for cat, retailers in skipped_list.items():
        log(f"  - {cat} (from: {', '.join(retailers)})")
else:
    log("✅ No skipped categories due to missing retailer coverage.")
